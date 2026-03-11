# =============================================================================
# PBTA Gene Correlation Analysis + Figure Generator
# =============================================================================
# This script computes pairwise Pearson correlations between a target gene of
# interest (e.g. ST8SIA2, ST8SIA4) and all other genes in the PBTA bulk
# RNA-seq expression matrix (RSEM TPM), then generates publication figures.
#
# Workflow:
#   1. Load the PBTA stranded TPM expression matrix
#   2. Load clinical metadata and filter to pHGG/DIPG tumor types
#   3. Parse heterogeneous gene ID formats (ENSG|SYMBOL, ENSG_SYMBOL, etc.)
#   4. Apply log2(TPM + 1) transformation to reduce skew before correlating
#   5. Identify the target gene by ENSG ID or gene symbol (highest-variance
#      isoform selected automatically when multiple entries exist)
#   6. Compute Pearson correlations + p-values across tumor-filtered samples
#   7. Apply Benjamini-Hochberg FDR correction for multiple testing
#   8. Deduplicate multi-ENSG symbols, keeping the highest |r| entry
#   9. Save results as: TSV (always), Excel with heatmap (if openpyxl available),
#      CSV fallback
#  10. Generate figures:
#        - Volcano plot (r vs -log10 FDR, top hits labeled)
#        - Lollipop chart (top N positive + negative genes)
#
# Inputs:  PBTA TPM matrix (.tsv.gz) + clinical metadata TSV, configured below
# Outputs: Ranked correlation table + Excel heatmap + PDF/PNG figures in OUT_DIR
# =============================================================================

import difflib
import re
import warnings
from pathlib import Path
from typing import Optional

import matplotlib
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.lines import Line2D
from scipy import stats
from statsmodels.stats.multitest import multipletests

warnings.filterwarnings("ignore")

# =========================
# CONFIG
# =========================
PBTA_PATH = Path(
    "/Users/fudhailsayed/prololab/datasets/PBTA/pbta-gene-expression-rsem-tpm.stranded.tsv.gz"
)
CLINICAL_PATH = Path(
    "/Users/fudhailsayed/prololab/datasets/PBTA/pbta_all_clinical_data0.tsv"
)

# EDIT THIS TO YOUR GENE OF INTEREST
TARGET_ENSG = None  # e.g. "ENSG00000143557" (takes priority if set)
TARGET_SYMBOL = "ST8SIA2"  # e.g. "ST8SIA2"

OUT_DIR = Path(
    "/Users/fudhailsayed/Desktop/prololab/bioinformatic_analysis/bulk_rnaseq/gene_correlation"
)
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Output files
OUT_EXCEL = OUT_DIR / f"{TARGET_SYMBOL}_PBTA_pHGG_correlations.xlsx"
OUT_TSV = OUT_DIR / f"{TARGET_SYMBOL}_PBTA_pHGG_correlations.tsv"
OUT_CSV = OUT_DIR / f"{TARGET_SYMBOL}_PBTA_pHGG_correlations.csv"

# Analysis parameters
CORRELATION_METHOD = "pearson"
LOG2_TRANSFORM = True  # Recommended: log2(TPM + 1) before correlating
FDR_ALPHA = 0.05  # Significance threshold after BH correction

# Tumor type filter — matches on CANCER_TYPE_DETAILED (substring, case-insensitive)
# Captures: pHGG (IDH-mutant, IDH-wildtype, H3-altered), DIPG/DMG, pontine glioma
TUMOR_TYPE_PATTERNS = [
    "high-grade glioma",
    "diffuse midline glioma",
    "diffuse intrinsic pontine glioma",
    "brainstem glioma",
]

# Figure parameters
TOP_N_LABEL = 15  # genes to label on volcano plot
TOP_N_LOLLIPOP = 20  # genes per side on lollipop chart
FIG_DPI = 300  # output resolution (300 = publication quality)

# Publication color palette
C_POS = "#C0392B"  # significant positive — deep red
C_NEG = "#2471A3"  # significant negative — deep blue
C_NS = "#BDC3C7"  # not significant — light grey

matplotlib.rcParams.update(
    {
        "font.family": "DejaVu Sans",
        "axes.spines.top": False,
        "axes.spines.right": False,
        "axes.linewidth": 0.8,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "axes.labelsize": 11,
        "axes.titlesize": 12,
        "axes.titleweight": "bold",
    }
)


# =========================
# CHECK EXCEL SUPPORT
# =========================
def get_excel_engine():
    try:
        import openpyxl

        return "openpyxl"
    except ImportError:
        pass
    try:
        import xlsxwriter

        return "xlsxwriter"
    except ImportError:
        pass
    return None


EXCEL_ENGINE = get_excel_engine()
if EXCEL_ENGINE:
    print(f"Excel engine available: {EXCEL_ENGINE}")
else:
    print("⚠️  No Excel engine found (openpyxl or xlsxwriter) — will save as CSV")


# =========================
# HELPERS
# =========================
ENSG_RE = re.compile(r"(ENSG\d+)(?:\.\d+)?")


def parse_gene_id(s: str):
    """Parse heterogeneous PBTA gene ID formats into (ENSG, symbol) tuples."""
    if pd.isna(s):
        return (None, None)
    s = str(s).strip()
    if "|" in s:
        left, right = s.split("|", 1)
        m = ENSG_RE.search(left)
        ensg = m.group(1) if m else None
        symbol = right.strip() if right.strip() else None
        return (ensg, symbol)
    if "_" in s:
        left, right = s.split("_", 1)
        m = ENSG_RE.search(left)
        ensg = m.group(1) if m else None
        symbol = right.strip() if right.strip() else None
        return (ensg, symbol)
    parts = s.split()
    if len(parts) >= 2:
        m = ENSG_RE.search(parts[0])
        ensg = m.group(1) if m else None
        symbol = parts[1].strip() if parts[1].strip() else None
        return (ensg, symbol)
    m = ENSG_RE.search(s)
    ensg = m.group(1) if m else None
    return (ensg, None)


def pick_target_index(expr_df, target_ensg_input, target_symbol_input):
    """Identify the best ENSG row for the target gene."""
    if target_ensg_input:
        if target_ensg_input not in expr_df.index:
            close = difflib.get_close_matches(
                target_ensg_input, expr_df.index.tolist(), n=10, cutoff=0.6
            )
            raise KeyError(
                f"TARGET_ENSG '{target_ensg_input}' not found.\nClosest: {close}"
            )
        return target_ensg_input
    if target_symbol_input:
        hits = expr_df.loc[
            expr_df["gene_symbol"].str.upper() == target_symbol_input.upper(), :
        ]
        if hits.empty:
            symbols = expr_df["gene_symbol"].dropna().unique().tolist()
            close = difflib.get_close_matches(
                target_symbol_input.upper(),
                [x.upper() for x in symbols],
                n=10,
                cutoff=0.6,
            )
            raise KeyError(
                f"TARGET_SYMBOL '{target_symbol_input}' not found.\nClosest: {close}"
            )
        sample_cols = [c for c in expr_df.columns if c not in ("ensg", "gene_symbol")]
        var = hits[sample_cols].var(axis=1, skipna=True)
        best_ensg = var.idxmax()
        print(
            f"Found {len(hits)} entries for '{target_symbol_input}', "
            f"selecting {best_ensg} (highest variance)"
        )
        return best_ensg
    raise ValueError("You must set either TARGET_ENSG or TARGET_SYMBOL.")


# =========================
# LOAD CLINICAL METADATA & FILTER SAMPLES
# =========================
print("=" * 70)
print(f"{TARGET_SYMBOL} Gene Correlation Analysis — PBTA pHGG/DIPG")
print("=" * 70)

print(f"\nLoading clinical metadata from:\n  {CLINICAL_PATH}")
if not CLINICAL_PATH.exists():
    raise FileNotFoundError(f"Clinical file not found at: {CLINICAL_PATH}")

clin = pd.read_csv(CLINICAL_PATH, sep="\t", low_memory=False, on_bad_lines="skip")
print(f"✓ Loaded: {len(clin):,} clinical records, {len(clin.columns)} columns")

# Identify sample ID column
sample_id_col = None
for candidate in [
    "SPECIMEN_ID",
    "Sample ID",
    "Sample_ID",
    "SAMPLE_ID",
    "sample_id",
    "Kids_First_Biospecimen_ID",
]:
    if candidate in clin.columns:
        sample_id_col = candidate
        break
if sample_id_col is None:
    raise ValueError(
        f"Could not find sample ID column in clinical file.\n"
        f"Available columns: {clin.columns.tolist()}"
    )
print(f"✓ Sample ID column: '{sample_id_col}'")

# Identify tumor type column
tumor_col = None
for candidate in [
    "CANCER_TYPE_DETAILED",
    "cancer_type_detailed",
    "CANCER_TYPE",
    "cancer_type",
]:
    if candidate in clin.columns:
        tumor_col = candidate
        break
if tumor_col is None:
    raise ValueError(
        f"Could not find tumor type column.\nAvailable: {clin.columns.tolist()}"
    )
print(f"✓ Tumor type column: '{tumor_col}'")

# Filter to pHGG + DIPG samples
pattern = "|".join(TUMOR_TYPE_PATTERNS)
phgg_mask = clin[tumor_col].str.contains(pattern, case=False, na=False)
clin_phgg = clin[phgg_mask].copy()

# SPECIMEN_ID contains semicolon-separated BS_ IDs — split and collect all
raw_ids = clin_phgg[sample_id_col].dropna().astype(str)
phgg_sample_ids = set()
for val in raw_ids:
    for sid in val.split(";"):
        sid = sid.strip()
        if sid:
            phgg_sample_ids.add(sid)

print(f"\n📋 Tumor type breakdown (pHGG/DIPG filter):")
type_counts = clin_phgg[tumor_col].value_counts()
for tumor_type, count in type_counts.items():
    print(f"   {count:>4}  {tumor_type}")
print(f"\n✓ Total pHGG/DIPG clinical records: {len(clin_phgg):,}")
print(f"✓ Unique sample IDs to keep: {len(phgg_sample_ids):,}")


# =========================
# LOAD EXPRESSION MATRIX
# =========================
print(f"\nLoading PBTA expression matrix from:\n  {PBTA_PATH}")
if not PBTA_PATH.exists():
    raise FileNotFoundError(f"PBTA file not found at: {PBTA_PATH}")

expr = pd.read_csv(PBTA_PATH, sep="\t", low_memory=False)
print(f"✓ Loaded: {expr.shape[0]:,} genes × {expr.shape[1]:,} columns")

# Parse gene IDs
gene_col_candidates = ["gene_id", "gene", "Gene", "ensembl_gene_id", "Ensembl_ID"]
gene_col = next((c for c in gene_col_candidates if c in expr.columns), expr.columns[0])

print(f"\nParsing gene identifiers (column: '{gene_col}')...")
parsed = expr[gene_col].astype(str).apply(parse_gene_id)
expr["ensg"] = parsed.apply(lambda x: x[0])
expr["gene_symbol"] = parsed.apply(lambda x: x[1])
expr = expr.dropna(subset=["ensg"]).copy()
expr = expr.set_index("ensg", drop=False)
expr = expr[~expr.index.duplicated(keep="first")].copy()
print(f"✓ Processed: {len(expr):,} genes")

# Intersect with pHGG samples
meta_cols = {gene_col, "ensg", "gene_symbol"}
all_samples = [c for c in expr.columns if c not in meta_cols]
phgg_samples = [s for s in all_samples if s in phgg_sample_ids]

if len(phgg_samples) == 0:
    raise ValueError(
        f"No pHGG sample IDs matched expression matrix columns.\n"
        f"Example clinical IDs: {list(phgg_sample_ids)[:5]}\n"
        f"Example expression columns: {all_samples[:5]}"
    )
print(f"✓ pHGG samples matched in expression matrix: {len(phgg_samples):,}")

expr_numeric = expr[phgg_samples].apply(pd.to_numeric, errors="coerce")
expr_numeric = expr_numeric.dropna(axis=0, how="all")


# =========================
# LOG2 TRANSFORM
# =========================
if LOG2_TRANSFORM:
    print(f"\nApplying log2(TPM + 1) transformation...")
    expr_numeric = np.log2(expr_numeric + 1)
    print(f"✓ Log2-transformed")


# =========================
# SELECT TARGET GENE
# =========================
print(f"\nSelecting target gene...")
target_ensg_key = pick_target_index(
    expr_df=pd.concat([expr[["ensg", "gene_symbol"]], expr_numeric], axis=1),
    target_ensg_input=TARGET_ENSG,
    target_symbol_input=TARGET_SYMBOL,
)

target_symbol_resolved = expr.loc[target_ensg_key, "gene_symbol"]
print(f"✓ Target: {target_symbol_resolved} ({target_ensg_key})")

st8_vec = expr_numeric.loc[target_ensg_key].copy()
common_samples = expr_numeric.columns.intersection(st8_vec.index)
expr_aligned = expr_numeric[common_samples]
st8_aligned = st8_vec[common_samples].values
print(f"✓ Aligned samples: {len(common_samples):,}")

gene_var = expr_aligned.var(axis=1, skipna=True)
expr_aligned = expr_aligned.loc[gene_var > 0]
print(f"✓ Genes after zero-variance filter: {len(expr_aligned):,}")


# =========================
# COMPUTE CORRELATIONS + P-VALUES
# =========================
print(f"\nComputing {CORRELATION_METHOD.capitalize()} correlations + p-values...")
print(f"  (This may take a moment for {len(expr_aligned):,} genes...)")

corr_vals, pval_vals = [], []
for gene_ensg in expr_aligned.index:
    if gene_ensg == target_ensg_key:
        corr_vals.append(np.nan)
        pval_vals.append(np.nan)
        continue
    r, p = stats.pearsonr(st8_aligned, expr_aligned.loc[gene_ensg].values)
    corr_vals.append(r)
    pval_vals.append(p)

print(f"✓ Correlations computed")

res = pd.DataFrame(
    {
        "ensg": expr_aligned.index,
        "correlation": corr_vals,
        "pvalue": pval_vals,
    }
)
res["gene_symbol"] = res["ensg"].map(expr["gene_symbol"])
res = res[res["ensg"] != target_ensg_key].copy()
res = res.dropna(subset=["correlation", "pvalue"])
res["gene_symbol"] = res["gene_symbol"].fillna(res["ensg"])


# =========================
# BENJAMINI-HOCHBERG FDR CORRECTION
# =========================
print(f"\nApplying Benjamini-Hochberg FDR correction (alpha={FDR_ALPHA})...")
_, fdr_vals, _, _ = multipletests(
    res["pvalue"].values, alpha=FDR_ALPHA, method="fdr_bh"
)
res["FDR"] = fdr_vals
res["significant"] = res["FDR"] < FDR_ALPHA
print(
    f"✓ Significant genes (FDR < {FDR_ALPHA}): {res['significant'].sum():,} of {len(res):,}"
)


# =========================
# SORT & DEDUPLICATE
# =========================
res = res.sort_values("correlation", ascending=False).reset_index(drop=True)

duplicate_symbols = res[res["gene_symbol"].duplicated(keep=False)][
    "gene_symbol"
].unique()
if len(duplicate_symbols) > 0:
    print(f"  Deduplicating {len(duplicate_symbols):,} gene symbols...")
    res["abs_corr"] = res["correlation"].abs()
    res = res.sort_values("abs_corr", ascending=False)
    res = res.drop_duplicates(subset=["gene_symbol"], keep="first").copy()
    res = res.drop(columns=["abs_corr"])
    res = res.sort_values("correlation", ascending=False).reset_index(drop=True)

res.insert(0, "rank", range(1, len(res) + 1))
res = res[
    ["rank", "gene_symbol", "correlation", "pvalue", "FDR", "significant", "ensg"]
]
print(f"✓ Final results: {len(res):,} genes")


# =========================
# SAVE OUTPUTS
# =========================
print("\n" + "=" * 70)
print("Saving results...")
print("=" * 70)

res.to_csv(OUT_TSV, sep="\t", index=False)
print(f"✓ TSV saved: {OUT_TSV.name}")

if EXCEL_ENGINE:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        print(f"  Creating Excel with heatmap visualization...")
        wb = Workbook()

        # ============= HEATMAP SHEET — Top 15 =============
        ws_heatmap = wb.active
        ws_heatmap.title = "Top_15_Heatmap"
        top15 = res.head(15)[["rank", "gene_symbol", "correlation", "FDR"]].copy()

        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")

        ws_heatmap["A1"] = (
            f"Top 15 Genes Positively Correlated with {target_symbol_resolved} "
            f"(pHGG/DIPG, log2 TPM, BH-FDR)"
        )
        ws_heatmap.merge_cells("A1:D1")
        ws_heatmap["A1"].font = Font(bold=True, size=13, name="Arial")
        ws_heatmap["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_heatmap.row_dimensions[1].height = 28

        for col_letter, label in zip(
            ["A", "B", "C", "D"], ["Rank", "Gene Symbol", "Pearson r", "BH-FDR"]
        ):
            cell = ws_heatmap[f"{col_letter}2"]
            cell.value = label
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        min_corr = top15["correlation"].min()
        max_corr = top15["correlation"].max()

        for i, (_, row) in enumerate(top15.iterrows(), start=3):
            corr = row["correlation"]
            fdr = row["FDR"]
            normalized = (
                (corr - min_corr) / (max_corr - min_corr)
                if max_corr != min_corr
                else 1.0
            )

            if normalized > 0.66:
                intensity = int((normalized - 0.66) / 0.34 * 255)
                color = f"FF{255-intensity:02X}{255-intensity:02X}"
            elif normalized > 0.33:
                intensity = int((normalized - 0.33) / 0.33 * 255)
                color = f"FF{165+intensity//3:02X}00"
            else:
                intensity = int(normalized / 0.33 * 255)
                color = f"FF{255:02X}{200+intensity//5:02X}"

            ws_heatmap[f"A{i}"] = row["rank"]
            ws_heatmap[f"A{i}"].alignment = Alignment(horizontal="center")
            ws_heatmap[f"A{i}"].font = Font(name="Arial", size=10)

            ws_heatmap[f"B{i}"] = row["gene_symbol"]
            ws_heatmap[f"B{i}"].font = Font(name="Arial", size=10, bold=True)
            ws_heatmap[f"B{i}"].alignment = Alignment(horizontal="left")

            ws_heatmap[f"C{i}"] = corr
            ws_heatmap[f"C{i}"].number_format = "0.0000"
            ws_heatmap[f"C{i}"].font = Font(
                name="Arial", size=10, bold=True, color="000000"
            )
            ws_heatmap[f"C{i}"].fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws_heatmap[f"C{i}"].alignment = Alignment(horizontal="center")

            ws_heatmap[f"D{i}"] = fdr
            ws_heatmap[f"D{i}"].number_format = "0.00E+00"
            ws_heatmap[f"D{i}"].font = Font(
                name="Arial", size=10, color="FF0000" if fdr >= FDR_ALPHA else "000000"
            )
            ws_heatmap[f"D{i}"].alignment = Alignment(horizontal="center")

        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row_i in range(2, len(top15) + 3):
            for col in ["A", "B", "C", "D"]:
                ws_heatmap[f"{col}{row_i}"].border = thin

        ws_heatmap.column_dimensions["A"].width = 8
        ws_heatmap.column_dimensions["B"].width = 22
        ws_heatmap.column_dimensions["C"].width = 14
        ws_heatmap.column_dimensions["D"].width = 14

        # ============= ALL CORRELATIONS SHEET =============
        ws_corr = wb.create_sheet("All_Correlations")
        headers = [
            "Rank",
            "Gene Symbol",
            "Pearson r",
            "P-value",
            "BH-FDR",
            "Significant",
            "ENSG",
        ]
        col_widths = [8, 22, 12, 14, 14, 13, 20]

        for col_i, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws_corr.cell(row=1, column=col_i, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws_corr.column_dimensions[get_column_letter(col_i)].width = width

        for row_i, (_, row) in enumerate(res.iterrows(), start=2):
            ws_corr.cell(row=row_i, column=1, value=row["rank"])
            ws_corr.cell(row=row_i, column=2, value=row["gene_symbol"])
            c = ws_corr.cell(row=row_i, column=3, value=row["correlation"])
            c.number_format = "0.0000"
            p = ws_corr.cell(row=row_i, column=4, value=row["pvalue"])
            p.number_format = "0.00E+00"
            f = ws_corr.cell(row=row_i, column=5, value=row["FDR"])
            f.number_format = "0.00E+00"
            ws_corr.cell(row=row_i, column=6, value=str(row["significant"]))
            ws_corr.cell(row=row_i, column=7, value=row["ensg"])

        # ============= SUMMARY SHEET =============
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Target Gene", target_symbol_resolved],
            ["Target ENSG", target_ensg_key],
            ["Tumor Filter", "pHGG + DIPG (CANCER_TYPE_DETAILED)"],
            ["Log2 Transform", f"log2(TPM + 1): {LOG2_TRANSFORM}"],
            ["Correlation Method", CORRELATION_METHOD.capitalize()],
            ["Multiple Testing", f"Benjamini-Hochberg FDR (alpha={FDR_ALPHA})"],
            ["pHGG Samples Used", len(phgg_samples)],
            ["Total Genes Analyzed", len(res)],
            ["Significant (FDR<0.05)", int(res["significant"].sum())],
            ["Strongest Positive r", res["correlation"].max()],
            ["Strongest Negative r", res["correlation"].min()],
            ["Mean Correlation", round(res["correlation"].mean(), 4)],
            ["Date Generated", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        for row_i, (metric, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=row_i, column=1, value=metric)
            ws_summary.cell(row=row_i, column=2, value=value)
            if row_i == 1:
                for col_i in [1, 2]:
                    cell = ws_summary.cell(row=1, column=col_i)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial")

        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 35

        wb.save(OUT_EXCEL)
        print(f"✓ Excel with heatmap saved: {OUT_EXCEL.name}")

    except Exception as e:
        print(f"⚠️  Excel creation failed: {e}")
        res.to_csv(OUT_CSV, index=False)
        print(f"✓ CSV fallback saved: {OUT_CSV.name}")
else:
    res.to_csv(OUT_CSV, index=False)
    print(f"✓ CSV saved: {OUT_CSV.name}")


# =========================
# PRINT SUMMARY
# =========================
print("\n" + "=" * 70)
print("ANALYSIS SUMMARY")
print("=" * 70)
print(f"\nTarget:         {target_symbol_resolved} ({target_ensg_key})")
print(f"Tumor filter:   pHGG + DIPG")
print(f"Transform:      log2(TPM + 1)")
print(f"Samples:        {len(phgg_samples):,}")
print(f"Genes:          {len(res):,}")
print(f"\nStrongest positive:  r = {res['correlation'].max():.4f}")
print(f"Strongest negative:  r = {res['correlation'].min():.4f}")
print(f"\nSignificant (FDR < {FDR_ALPHA}):  {int(res['significant'].sum()):,} genes")

print("\n🔺 TOP 15 POSITIVELY CORRELATED GENES (pHGG/DIPG):")
print("-" * 70)
display = res.head(15)[["rank", "gene_symbol", "correlation", "FDR"]].copy()
display["correlation"] = display["correlation"].apply(lambda x: f"{x:+.4f}")
display["FDR"] = display["FDR"].apply(lambda x: f"{x:.2e}")
print(display.to_string(index=False))


# =============================================================================
# FIGURE GENERATION
# =============================================================================
print("\n" + "=" * 70)
print("Generating publication figures...")
print("=" * 70)

# Shared prep — color categories and label gene sets
res["-log10FDR"] = -np.log10(res["FDR"].clip(lower=1e-300))


def assign_color_cat(row):
    if not row["significant"]:
        return "ns"
    return "pos" if row["correlation"] > 0 else "neg"


res["color_cat"] = res.apply(assign_color_cat, axis=1)

top_pos_label = res[res["color_cat"] == "pos"].nlargest(TOP_N_LABEL, "correlation")
top_neg_label = res[res["color_cat"] == "neg"].nsmallest(TOP_N_LABEL, "correlation")
label_genes = pd.concat([top_pos_label, top_neg_label])


# -----------------------------------------------------------------------
# FIGURE 1: VOLCANO PLOT
# -----------------------------------------------------------------------
print("\nGenerating Figure 1: Volcano plot...")

fig1, ax = plt.subplots(figsize=(8, 6))

# Non-significant background
ns_dots = res[res["color_cat"] == "ns"]
ax.scatter(
    ns_dots["correlation"],
    ns_dots["-log10FDR"],
    c=C_NS,
    s=8,
    alpha=0.4,
    linewidths=0,
    rasterized=True,
    zorder=1,
)

# Significant dots
for cat, color in [("pos", C_POS), ("neg", C_NEG)]:
    sub = res[res["color_cat"] == cat]
    ax.scatter(
        sub["correlation"],
        sub["-log10FDR"],
        c=color,
        s=12,
        alpha=0.7,
        linewidths=0,
        rasterized=True,
        zorder=2,
    )

# FDR threshold + zero lines
fdr_line = -np.log10(FDR_ALPHA)
ax.axhline(fdr_line, color="black", linestyle="--", linewidth=0.8, alpha=0.6, zorder=3)
ax.axvline(0, color="black", linestyle="-", linewidth=0.5, alpha=0.3, zorder=3)
ax.text(
    res["correlation"].min(),
    fdr_line + 0.3,
    f"FDR = {FDR_ALPHA}",
    fontsize=8,
    color="black",
    alpha=0.7,
)

# Gene labels
try:
    from adjustText import adjust_text

    texts = []
    for _, row in label_genes.iterrows():
        t = ax.text(
            row["correlation"],
            row["-log10FDR"],
            row["gene_symbol"],
            fontsize=6.5,
            color="black",
            fontweight="bold",
            ha="center",
            va="bottom",
        )
        texts.append(t)
    adjust_text(
        texts,
        ax=ax,
        arrowprops=dict(arrowstyle="-", color="gray", lw=0.5),
        expand_points=(1.5, 1.5),
    )
except ImportError:
    for _, row in label_genes.iterrows():
        ax.annotate(
            row["gene_symbol"],
            xy=(row["correlation"], row["-log10FDR"]),
            xytext=(3, 3),
            textcoords="offset points",
            fontsize=6.5,
            fontweight="bold",
            arrowprops=dict(arrowstyle="-", color="gray", lw=0.4),
        )

ax.set_xlabel(f"Pearson r (correlation with {TARGET_SYMBOL})", labelpad=8)
ax.set_ylabel("-log$_{10}$(BH-FDR)", labelpad=8)
ax.set_title(
    f"{TARGET_SYMBOL} Gene Correlations — PBTA pHGG/DIPG\n"
    f"(n={len(res):,} genes, log$_2$ TPM)",
    pad=12,
)

legend_elements = [
    mpatches.Patch(
        facecolor=C_POS, label=f"Sig. positive ({(res['color_cat']=='pos').sum():,})"
    ),
    mpatches.Patch(
        facecolor=C_NEG, label=f"Sig. negative ({(res['color_cat']=='neg').sum():,})"
    ),
    mpatches.Patch(
        facecolor=C_NS, label=f"Not significant ({(res['color_cat']=='ns').sum():,})"
    ),
]
ax.legend(handles=legend_elements, fontsize=8, frameon=False, loc="upper left")

plt.tight_layout()
out_volcano_pdf = OUT_DIR / f"{TARGET_SYMBOL}_volcano.pdf"
out_volcano_png = OUT_DIR / f"{TARGET_SYMBOL}_volcano.png"
fig1.savefig(out_volcano_pdf, dpi=FIG_DPI, bbox_inches="tight")
fig1.savefig(out_volcano_png, dpi=FIG_DPI, bbox_inches="tight")
print(f"✓ Saved: {out_volcano_pdf.name}")
plt.close(fig1)


# -----------------------------------------------------------------------
# FIGURE 2: LOLLIPOP CHART
# -----------------------------------------------------------------------
print("Generating Figure 2: Lollipop chart...")

top_pos_lol = res[res["correlation"] > 0].nlargest(TOP_N_LOLLIPOP, "correlation")
top_neg_lol = res[res["correlation"] < 0].nsmallest(TOP_N_LOLLIPOP, "correlation")
lol_df = (
    pd.concat([top_neg_lol, top_pos_lol])
    .sort_values("correlation")
    .reset_index(drop=True)
)

fig2, ax = plt.subplots(figsize=(7, 10))

colors = [C_NEG if r < 0 else C_POS for r in lol_df["correlation"]]
y_pos = np.arange(len(lol_df))

# Stems and dots
ax.hlines(y_pos, 0, lol_df["correlation"], color=colors, linewidth=1.2, alpha=0.7)
ax.scatter(
    lol_df["correlation"],
    y_pos,
    color=colors,
    s=55,
    zorder=5,
    edgecolors="white",
    linewidths=0.5,
)

# Black ring on significant genes
for i, (_, row) in enumerate(lol_df.iterrows()):
    if row["significant"]:
        ax.scatter(
            row["correlation"],
            i,
            color=colors[i],
            s=55,
            zorder=6,
            edgecolors="black",
            linewidths=0.8,
        )

# Gene labels — offset far enough from dot so text doesn't overlap
for i, (_, row) in enumerate(lol_df.iterrows()):
    ha = "left" if row["correlation"] > 0 else "right"
    pad = 0.025 if row["correlation"] > 0 else -0.025
    ax.text(
        row["correlation"] + pad,
        i,
        row["gene_symbol"],
        va="center",
        ha=ha,
        fontsize=7.5,
        fontweight="bold",
    )

ax.axvline(0, color="black", linewidth=0.8, alpha=0.5)
ax.set_yticks([])
ax.set_xlabel(f"Pearson r (correlation with {TARGET_SYMBOL})", labelpad=8)
ax.set_title(
    f"Top {TOP_N_LOLLIPOP} Positively & Negatively Correlated Genes\n"
    f"{TARGET_SYMBOL} — PBTA pHGG/DIPG (log$_2$ TPM)",
    pad=12,
)

max_r = lol_df["correlation"].abs().max()
ax.set_xlim(-max_r * 1.6, max_r * 1.6)
ax.set_ylim(-0.8, len(lol_df) - 0.2)

legend_elements = [
    mpatches.Patch(facecolor=C_POS, label="Positive correlation"),
    mpatches.Patch(facecolor=C_NEG, label="Negative correlation"),
    Line2D(
        [0],
        [0],
        marker="o",
        color="w",
        markerfacecolor="gray",
        markeredgecolor="black",
        markersize=7,
        label=f"FDR < {FDR_ALPHA}",
    ),
]
ax.legend(handles=legend_elements, fontsize=8, frameon=False, loc="lower right")

plt.tight_layout()
out_lollipop_pdf = OUT_DIR / f"{TARGET_SYMBOL}_lollipop.pdf"
out_lollipop_png = OUT_DIR / f"{TARGET_SYMBOL}_lollipop.png"
fig2.savefig(out_lollipop_pdf, dpi=FIG_DPI, bbox_inches="tight")
fig2.savefig(out_lollipop_png, dpi=FIG_DPI, bbox_inches="tight")
print(f"✓ Saved: {out_lollipop_pdf.name}")
plt.close(fig2)


# =========================
# FINAL SUMMARY
# =========================
print("\n" + "=" * 70)
print("✅ ANALYSIS COMPLETE")
print("=" * 70)
print(f"\n📊 {OUT_EXCEL.name}")
print(f"📄 {OUT_TSV.name}")
print(f"🖼️  {out_volcano_pdf.name}  +  {out_volcano_png.name}")
print(f"🖼️  {out_lollipop_pdf.name}  +  {out_lollipop_png.name}")
print(f"\nAll outputs saved to:\n  {OUT_DIR}")
print()
